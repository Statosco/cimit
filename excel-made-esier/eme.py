import os
import openpyxl
import ast



def updateFile(path):
    print('-'*70)
    data_type = input("What type of data are you trying to write? --> ")
    print('_'*70)
    print(f"Writing '{data_type}' data:")

    try:
        with open(path, "a") as exFile:
            print('-'*70)
            print("Paste or write your data below (Type STOP(uppercased) and press Enter to finish):")
            print('_'*70)
            while True:
                line = input()
                if line == "STOP":
                    break
                exFile.write(line + '\n')
            print('-'*70)
            print("Data has been written to the file.")
            print('-'*70)
    except Exception as e:
        print('-'*70)
        print(f"An error occurred:\n{e}")
        print('_'*70)


        
def createFile():
    while True:
        print('-'*70)
        files = input("What Data File Are You Trying To Create?\nWe Support Various Data Types Inculding\n\tCSV 'Comma Separated Values'\n\tTXT 'for word files or text files'\n\tXLSX 'For excel spread sheet'--> ").lower()
        print('_'*70)
        if files == "csv" or files == "txt":
            print('-'*70)
            name = input(f"which name should we give your {files} file? ")
            print('_'*70)
            if name.endswith(files):
                file = name
                checkForFile(file)
                print()
            else:
                file =(f"{name}.{files}")
                checkForFile(file)
            try:
                with open (file, 'w') as newFile:
                    newFile.write(updateFile(file))
            except Exception as e:
                print('-'*70)
                print(f"an error ocured while creating file\n{e}")
                print('_'*70)
                exit()
        elif files == "xlsx" or files == "excel":
            print('-'*70)
            fileName = input("which name should we give your excel file? --> ")
            print('_'*70)
            if fileName.endswith(files):
                exelFile = fileName
                checkForFile(exelFile)
                exelSheet(exelFile)
            else:

                fileWithE = (f"{fileName}.xlsx")
                checkForFile(fileWithE)
                exelSheet(fileWithE)
                    
        else:
            print('-'*70)
            print(f"the file format {files} does not exist in our scope")
            print('_'*70)
            
def exelSheet(path):
    print('-'*70)
    writeOption = input("would you like to write individual cells? (y/n) --> ").lower()
    print('_'*70)
    if writeOption == 'yes' or writeOption == 'y':
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Add data to the spreadsheet
        print('-'*70)
        Erows = int(input("how many rows --> "))
        cols = int(input("how many column --> "))
        print('_'*70)
        for row_num in range(1, Erows + 1):
            for col_num in range(1, cols + 1):
                print('-'*50)
                cell_value = input(f"Enter value for cell {openpyxl.utils.get_column_letter(col_num)}{row_num}: ")
                print('-'*50)
                sheet.cell(row=row_num, column=col_num, value=cell_value)
        print("file writen succesfully!")
        if '.xlsx' in path:
            workbook.save(path)
        else:
            workbook.save(f'{path}.xlsx')
            try:
                if os.path.exists(path):
                    os.startfile(path)
            except Exception as e:
                print(f"An error occurred while opening the directory: {e}")
            exit()
    elif writeOption == "no" or writeOption == "n":
        print('-'*70)
        formatOption = input("Are you trying to write the data of each rows individually \nOr are you trying to upload data from an external file ('1' or '2') --> ")
        print('_'*70)
        if formatOption == '1':

            wb = openpyxl.Workbook()
            sheet = wb.active
            while True:
                try:
                    print('-'*50)
                    data = int(input("how many rows do you want --> "))
                    print('_'*50)
                    break
                except ValueError:
                        print('-'*50)
                        print("rows has to be a number")
                        print('_'*50)

            for i in range(1, data + 1):
                print('-'*50)
                row = input(f"Write The Entire Row{i} --> ")
                print('_'*50)
                rows = row.split(',')
                sheet.append(rows)
            print("file written succesfully")
            if '.xlsx' in path:
                wb.save(path)
            else:
                wb.save(f'{path}.xlsx')
                try:
                    if os.path.exists(path):
                        os.startfile(path)
                except Exception as e:
                    exit()
        elif formatOption == '2':
            formatData()
    else:
        print('-'*50)
        print("option not in scope")
        print('_'*50)
        exelSheet(path)

def formatData():
    print('-'*70)
    filesPath = input("write file name and extension --> ")
    print('_'*70)
    filesPath = filesPath.replace('/',"\\")
    def isNestedLIst(data):
        try:
            parsedData = ast.literal_eval(data)
            if isinstance(parsedData, list) and all(isinstance(item, list) for item in parsedData) or(isinstance(parsedData, list)):
                return True
        except(ValueError, SyntaxError):
            pass
        return False

            
    if os.path.exists(filesPath):
        print('-'*50)
        exelFile = input("what is the name of your exel file in which datta will be written to --> ")
        checkForFile(exelFile)
        print('_'*50)
        try:
            with open(filesPath, 'r')as fp:
                contents = fp.read()
                if isNestedLIst(contents):
                    split_data = [item[0].split(',') for item in ast.literal_eval(contents)]
                    with open(filesPath, 'w') as fFile:
                        for row in split_data:
                            rowStr = ','.join(row)
                            fFile.write(f"{rowStr}\n")
                        wb = openpyxl.Workbook()
                        cells = wb.active
                        for rows in split_data:
                            cells.append(rows)
                    
    
                    if '.xlsx' in exelFile:
                        wb.save(exelFile)
                    else:
                        wb.save(f'{exelFile}.xlsx')
                
        except FileNotFoundError:
            print('-'*50)
            print("file was not foud")
            print('_'*50)
            formatData()
        except Exception as e:
            print('-'*50)
            print(f'an error ocured\n{e}\n')
            print('_'*50)
            return
    print('-'*50)
    dattaToRemove = input("Is there a character you would like to remove (yes/no)? ").lower()
    print('_'*50)
    if dattaToRemove in ['y', 'yes']:
        print('-'*50)
        target_characters = input("\nWhich characters do you want to remove? Enter the characters: ")
        print('_'*50)
        try:
            with open(filesPath, 'r') as data_file:
                content = data_file.read()

                for target_character in target_characters:
                    content = content.replace(target_character, '')

            with open(filesPath, 'w') as data_file:
                data_file.write(content)
                
                my_wb = openpyxl.Workbook()
                my_ws = my_wb.active

                content_lines = content.split('\n')
                for line in content_lines:
                    my_ws.append(line.split(','))
                if not exelFile.endswith('.xlsx'):
                    exelFile += '.xlsx'
                my_wb.save(exelFile)

                print('-'*50)
                print(f"Data has been written to '{exelFile}'")
                print('_'*50)
                exit()
                
        except Exception as e:
            print('-'*50)
            print(f"An error occurred: {e}")
            print('_'*50)
            return
    elif not dattaToRemove in ['y', 'yes']:
        try:
            with open(filesPath, 'r') as data_file:
                content = data_file.read()
            my_wb = openpyxl.Workbook()
            my_ws = my_wb.active

            content_lines = content.split('\n')
            for line in content_lines:
                my_ws.append(line.split(','))
            print('-'*50)
            if not exelFile.endswith('.xlsx'):
                exelFile += '.xlsx'
                my_wb.save(exelFile)
                print(f"Data has been written to '{exelFile}'")
                print('-'*50)
                exit()
            else:
                my_wb.save(exelFile)
        except Exception as e:
            print(e)
    else:
        print('-'*50)
        print("Something went wrong :(")
        print('_'*50)
        return


def checkForFile(file):
    if os.path.exists(file):
        print('-'*50)
        isChange = input("This File Exist Would like to Change? ").lower()
        print('-'*50)
        if isChange == "yes" or isChange == "y":
            try:
                print('-'*50)
                newName = input("Which Name Would Like To Give The File? ")
                print('_'*50)
                if os.path.exists(newName):
                    checkForFile(file)
                else:
                    _, ext = os.path.splitext(file)
                    newFileWithExt = newName + ext
                    if os.path.exists(newFileWithExt):
                        checkForFile(file)
                    else:
                        os.rename(file, newFileWithExt)
                        if ".xlsx" in newFileWithExt:
                            exelSheet(newFileWithExt)
                        else:
                            myFile = f"{newFileWithExt}.xlsx"
                            exelSheet(myFile)
                            
            except Exception as e:
                print('-'*50)
                print(f"An Error Ocured: \n'{e}'")
                print('_'*50)
                exit()
        else:
            print('-'*50)
            option = input("Do you want delete this file?(y/n) --> ").lower()
            print('_'*50)
            if option in ['y','yes']:
                os.remove(file)
            else:
                print('-'*50)
                print("Try a new file name")
                print('-'*50)
                exit()




def options():
    while True:
        print('-'*50)
        answer = input("What Are You Trying To Do Today?\n\tOption1: Create A file.\n\tOption2: Edit A File.\n\tOption3 Delete A File\n\tOr (Exit/q) To Quit: ").lower()
        print('_'*70)
        if answer in ['option1', 'create a file', '1']:
            createFile()
        
        elif answer in ['option2', 'edit a file', '2']:
            editFile()
        elif answer in ['option3', 'delete a file', '3']:
            deleteFile()
        elif answer in ['exit', 'q']:
            break
        else:
            print('-'*150)
            print(f"The Option/Keyword '{answer}' Is Not In Our Option's List Please Try agin\n")
            answer = input("What Are You Trying To Do Today?\n\tOption1: Create A file.\nOption2: Edit A File.\nOption3: Delete A File").lower()
            print('_'*150)

def deleteFile():
    print('-'*50)
    filePath = input("Which File Do You Want To Remove?\nWrite The File Path: ")
    print('-'*50)
    checkFile = os.path.exists(filePath)
    filePath = filePath.replace("/", "\\")
    if checkFile:
        print('-'*50)
        fileToDelete = input("this file will be deleted are you sure you want to delete? ").lower()
        print('_'*70)
        if fileToDelete == "y" or fileToDelete == "yes":
            os.remove(filePath)
        else:
            return
    else:
        print('-'*50)
        print(f"This File Path '{filePath}' Seems Not To Exists")
        print('_'*50)
        return


def printContents(fileName):
    if os.path.exists(fileName):
        try:
            with open (fileName, 'r')as file:
                pFile = file.read()
                print(pFile)
        except FileNotFoundError:
            print("this file does not exist or was removed")
        except Exception as e:
            print(f"an error ocured \n{e}")
    else:
        name = input("specify a full path")
        name=name.replace("/","\\")
        if os.path.exists(name):
            printContents(name)
        else:
            print("good bye")
            exit()
                



def editFile():
    print('-'*50)
    fileName = input("do you want to replace the data or edit a cpecific row?(r) for replace, (er) fo row, (ec ) for column:  ").lower()
    print('-'*50)


    if fileName == 'r':
        print('-'*50)
        myPath = input ("enter a file name ")
        print('-'*50)
        myPath= myPath.replace('/','\\')
        if os.path.exists(myPath):
            try:
                with open (myPath, 'w') as replaceFile:
                    replaceFile.write(updateFile(myPath))
            except Exception as e:
                print('-'*50)
                print(f"this file has an isue \n{e}")
                print('-'*50)
                return
        else:
            print('-'*50)
            ifNot = input("the file name was not found\nplease write the full path to it: ")
            print('-'*50)
            ifNot = ifNot.replace("/","\\")
            if os.path.exists(ifNot):
                try:
                    with open (myPath, 'w') as replaceFile:
                        replaceFile.write(updateFile(if_not))
                except Exception as e:
                    print('-'*50)
                    print(f"this file has an isue \n{e}")
                    print('-'*50)
                    return
            else:
                print('-'*50)
                if_not = ("file was not found \nEXIT? ").lower()
                print('-'*50)
                if if_not == 'y' or if_not == 'yes':
                    return
                else:
                    editFile()
    else:
        print('-'*50)
        print(f"this '{fileName}' is not in our option")
        print('-'*50)
        return


def main():
    
    print('-'*150)
    user = input("Hello! And welcome before we begin, I would like to ask you some questions so that I can know how best to process your data.\nWhat data is it intended for (coumpany name)-->> ").upper()
    print('_'*150)
    import pyfiglet
    pyfiglet.print_figlet(user,)
    paths = input("please enter the name of your exel file-->> ")
    if paths.endswith("xlsx"):
        checkForFile(paths)
    else:
        verify = f'{paths}.xlsx'
        checkForFile(verify)
    exelSheet(paths) 

main()